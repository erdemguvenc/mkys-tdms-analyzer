from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from analyzer.reconciliation.result import ReconciliationResult
from analyzer.reports.dashboard_summary import DashboardSummary

from .dashboard_writer import DashboardWriter

from .report_builder import ReportBuilder
from .worksheet_writer import WorksheetWriter


class ExcelReportBuilder(ReportBuilder):
    """
    Uzlaştırma sonucunu Excel dosyası olarak oluşturur.
    """

    def __init__(self) -> None:

        self._writer = WorksheetWriter()

        self._dashboard_writer = DashboardWriter()

    def build(
        self,
        result: ReconciliationResult,
        output_file: Path,
    ) -> None:

        workbook = Workbook()

        #
        # 0_Dashboard
        #
        dashboard = workbook.active

        dashboard.title = "0_Dashboard"

        summary = DashboardSummary.from_result(
            result,
        )

        self._dashboard_writer.write_dashboard(
            dashboard,
            summary,
        )

        #
        # 1_Özet
        #
        summary_sheet = workbook.create_sheet(
            "1_Özet",
        )

        self._writer.write_summary(
            summary_sheet,
            result,
        )

        #
        # 2_Giriş_Eşleşen
        #
        sheet = workbook.create_sheet(
            "2_Giriş_Eşleşen",
        )

        self._writer.write_movements(
            sheet,
            "Giriş Eşleşen",
            result.matched,
        )

        #
        # 3_MKYS_Eksik
        #
        sheet = workbook.create_sheet(
            "3_MKYS_Eksik",
        )

        self._writer.write_movements(
            sheet,
            "MKYS'de Bulunup TDMS'de Bulunmayan Girişler",
            result.missing_in_tdms,
        )

        #
        # 4_TDMS_Eksik
        #
        sheet = workbook.create_sheet(
            "4_TDMS_Eksik",
        )

        self._writer.write_movements(
            sheet,
            "TDMS'de Bulunup MKYS'de Bulunmayan Girişler",
            result.missing_in_mkys,
        )

        #
        # 5_Tutar_Farkları
        #
        sheet = workbook.create_sheet(
            "5_Tutar_Farkları",
        )

        self._writer.write_amount_differences(
            sheet,
            result.amount_differences,
        )

        #
        # 6_Tüketim_Farkları
        #
        sheet = workbook.create_sheet(
            "6_Tüketim_Farkları",
        )

        self._writer.write_consumption_differences(
            sheet,
            result.consumption_differences,
        )

        output_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(output_file)