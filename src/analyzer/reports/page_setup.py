from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins


DEFAULT_FREEZE_PANES = "A4"

DEFAULT_REPEAT_ROWS = "1:3"

DEFAULT_FIT_TO_WIDTH = 1

DEFAULT_FIT_TO_HEIGHT = 0


def prepare_worksheet(
    worksheet: Worksheet,
    *,
    header_row: int = 3,
) -> None:
    """
    Sayfaya kurumsal Excel düzenini uygular.
    """

    apply_freeze_panes(worksheet)
    apply_auto_filter(worksheet)
    apply_page_layout(worksheet)
    apply_print_settings(worksheet)


def apply_freeze_panes(
    worksheet: Worksheet,
) -> None:
    """
    Başlık satırlarını sabitler.
    """

    worksheet.freeze_panes = DEFAULT_FREEZE_PANES


def apply_auto_filter(
    worksheet: Worksheet,
) -> None:
    """
    Yazılmış tablo üzerinde otomatik filtreyi etkinleştirir.
    """

    if (
        worksheet.max_row >= 3
        and worksheet.max_column >= 1
    ):
        worksheet.auto_filter.ref = worksheet.dimensions


def apply_page_layout(
    worksheet: Worksheet,
) -> None:
    """
    Sayfa yönü ve kâğıt boyutu.
    """

    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4


def apply_print_settings(
    worksheet: Worksheet,
) -> None:

    worksheet.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )

    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.print_title_rows = "1:3"

    worksheet.sheet_view.showGridLines = True

    worksheet.print_options.gridLines = False

    worksheet.print_options.horizontalCentered = True

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.print_area = worksheet.dimensions


def apply_header_footer(
    worksheet: Worksheet,
    title: str,
) -> None:
    """
    Sayfa üst ve alt bilgilerini ayarlar.
    """

    worksheet.oddHeader.center.text = title
    worksheet.oddHeader.center.font = "Calibri,Bold"

    worksheet.oddFooter.right.text = "Sayfa &[Page]/&N"