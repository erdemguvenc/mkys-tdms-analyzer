from __future__ import annotations

from typing import cast

from openpyxl.cell import Cell
from openpyxl.chart import (
    BarChart,
    PieChart,
)
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.worksheet import Worksheet

from analyzer.reports.dashboard_summary import DashboardSummary
from analyzer.reports.executive_summary import (
    ExecutiveSummary,
    ExecutiveSummaryResult,
)
from analyzer.reports.report_status import ReportStatus

from .page_setup import prepare_worksheet
from .styles import (
    apply_kpi_card,
    apply_kpi_title,
    apply_kpi_value,
    apply_status_badge,
    apply_summary_critical,
    apply_summary_success,
    apply_summary_text,
    apply_summary_title,
    apply_summary_warning,
    apply_title,
)
from .theme import THEME_CHART_PRIMARY


class DashboardWriter:
    def write_dashboard(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        """
        Dashboard sayfasını oluşturur.
        """

        #
        # Başlık
        #
        self._write_dashboard_title(
            worksheet,
        )

        #
        # Executive Summary
        #
        summary_builder = ExecutiveSummary()

        result = summary_builder.build(
            summary,
        )

        self._write_executive_summary(
            worksheet,
            result,
        )

        self._write_status_badge(
            worksheet,
            result,
        )

        #
        # KPI Kartları
        #
        self._write_quality_progress(
            worksheet,
            summary,
        )

        self._write_kpis(
            worksheet,
            summary,
        )

        #
        # Dashboard Grafikleri
        #
        self._write_charts(
            worksheet,
            summary,
        )

        #
        # Sayfa hazırlığı
        #
        self._prepare_sheet(
            worksheet,
        )

    #
    # Bölümler
    #

    def _write_dashboard_title(
        self,
        worksheet: Worksheet,
    ) -> None:
        title = worksheet["A1"]
        title.value = "MKYS - TDMS Uzlaştırma Dashboard"

        apply_title(title)

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=12,
        )

    def _write_executive_summary(
        self,
        worksheet: Worksheet,
        result: ExecutiveSummaryResult,
    ) -> None:
        """
        Executive Summary bölümünü oluşturur.
        """

        #
        # Kutu
        #
        worksheet.merge_cells("A2:J2")
        worksheet.merge_cells("A3:J3")
        worksheet.merge_cells("A4:J4")
        worksheet.merge_cells("A5:J5")
        worksheet.merge_cells("A6:J6")

        #
        # Duruma göre kutuyu boya
        #
        if result.status is ReportStatus.GOOD:
            apply_summary_success(
                worksheet,
                first_row=2,
                last_row=6,
                first_column=1,
                last_column=10,
            )

        elif result.status is ReportStatus.WARNING:
            apply_summary_warning(
                worksheet,
                first_row=2,
                last_row=6,
                first_column=1,
                last_column=10,
            )

        else:
            apply_summary_critical(
                worksheet,
                first_row=2,
                last_row=6,
                first_column=1,
                last_column=10,
            )

        #
        # Başlık
        #
        title = worksheet["A2"]
        title.value = "Executive Summary"

        apply_summary_title(title)

        #
        # Metinler
        #
        cells = [
            worksheet["A3"],
            worksheet["A4"],
            worksheet["A5"],
            worksheet["A6"],
        ]

        for cell, line in zip(cells, result.lines):
            cell.value = line
            apply_summary_text(cell)

    def _write_status_badge(
        self,
        worksheet: Worksheet,
        result: ExecutiveSummaryResult,
    ) -> None:
        """
        Dashboard durum rozetini oluşturur.
        """

        badge = worksheet["K2"]

        if result.status is ReportStatus.GOOD:
            badge.value = "GOOD"

        elif result.status is ReportStatus.WARNING:
            badge.value = "WARNING"

        else:
            badge.value = "CRITICAL"

        apply_status_badge(
            badge,
            result.status,
        )

    def _write_kpi_card(
        self,
        worksheet: Worksheet,
        icon: str,
        title: str,
        value: int,
        row: int,
        column: int,
    ) -> None:
        """
        Tek bir KPI kartı oluşturur.
        """

        apply_kpi_card(
            worksheet,
            first_row=row,
            last_row=row + 2,
            first_column=column,
            last_column=column + 1,
        )

        worksheet.merge_cells(
            start_row=row,
            start_column=column,
            end_row=row,
            end_column=column + 1,
        )

        worksheet.merge_cells(
            start_row=row + 1,
            start_column=column,
            end_row=row + 2,
            end_column=column + 1,
        )

        title_cell = worksheet.cell(
            row=row,
            column=column,
        )

        title_cell = cast(Cell, title_cell)
        title_cell.value = f"{icon} {title}"

        apply_kpi_title(title_cell)

        value_cell = worksheet.cell(
            row=row + 1,
            column=column,
        )

        value_cell = cast(Cell, value_cell)
        value_cell.value = value

        apply_kpi_value(value_cell)

    def _write_kpis(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        """
        Dashboard KPI kartlarını oluşturur.
        """

        cards = [
            (
                "📊",
                "Toplam Kayıt",
                summary.total_records,
                8,
                1,
            ),
            (
                "✔",
                "Eşleşen Kayıt",
                summary.matched_records,
                8,
                4,
            ),
            (
                "⚠",
                "MKYS Eksik",
                summary.missing_in_mkys,
                8,
                7,
            ),
            (
                "⚠",
                "TDMS Eksik",
                summary.missing_in_tdms,
                8,
                10,
            ),
            (
                "💰",
                "Tutar Farkı",
                summary.amount_difference_count,
                12,
                1,
            ),
            (
                "📦",
                "Tüketim Farkı",
                summary.consumption_difference_count,
                12,
                4,
            ),
        ]

        for icon, title, value, row, column in cards:
            self._write_kpi_card(
                worksheet,
                icon,
                title,
                value,
                row,
                column,
            )

    def _write_quality_progress(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        """
        Dashboard kalite yüzdesini yazar.
        """

        worksheet["J7"] = "Kalite"

        quality = 0.0

        if summary.total_records:
            quality = summary.matched_records / summary.total_records

        cell = worksheet["K7"]
        cell.value = quality
        cell.number_format = "0%"

        worksheet.conditional_formatting.add(
            "K7",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=1,
                color=THEME_CHART_PRIMARY,
            ),
        )

    #
    # Grafikler
    #

    def _write_charts(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        self._write_match_pie_chart(
            worksheet,
            summary,
        )

        self._write_consumption_chart(
            worksheet,
            summary,
        )

    def _write_match_pie_chart(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        """
        Eşleşme durumunu gösteren pasta grafiği.
        """

        worksheet["N1"] = "Durum"
        worksheet["O1"] = "Kayıt"

        worksheet["N2"] = "Eşleşen"
        worksheet["O2"] = summary.matched_records

        worksheet["N3"] = "MKYS Eksik"
        worksheet["O3"] = summary.missing_in_mkys

        worksheet["N4"] = "TDMS Eksik"
        worksheet["O4"] = summary.missing_in_tdms

        chart = PieChart()

        chart.title = "Giriş Hareketleri"

        chart.height = 7
        chart.width = 14

        worksheet.add_chart(
            chart,
            "M2",
        )

    def _write_consumption_chart(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        """
        Tutar ve tüketim farklarını gösteren sütun grafiği.
        """

        worksheet["R1"] = "Kategori"
        worksheet["S1"] = "Kayıt"

        worksheet["R2"] = "Tutar Farkı"
        worksheet["S2"] = summary.amount_difference_count

        worksheet["R3"] = "Tüketim Farkı"
        worksheet["S3"] = summary.consumption_difference_count

        chart = BarChart()

        chart.type = "col"
        chart.style = 10

        chart.title = "MKYS - TDMS Fark Analizi"

        chart.y_axis.title = "Kayıt"
        chart.x_axis.title = "Kategori"

        chart.height = 7
        chart.width = 14

        worksheet.add_chart(
            chart,
            "M20",
        )

    #
    # Yardımcılar
    #

    def _prepare_sheet(
        self,
        worksheet: Worksheet,
    ) -> None:
        prepare_worksheet(
            worksheet,
        )
