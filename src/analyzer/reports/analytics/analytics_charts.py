from __future__ import annotations

from openpyxl.chart import (
    BarChart,
    LineChart,
    Reference,
)
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.worksheet import Worksheet

CHART_HEIGHT = 7
CHART_WIDTH = 14

TREND_CHART_POSITION = "J2"
SUPPLIER_CHART_POSITION = "J22"
WAREHOUSE_CHART_POSITION = "J42"
TOP_DIFFERENCE_CHART_POSITION = "J62"


class AnalyticsCharts:
    """
    Analytics sayfasındaki grafikleri oluşturur.
    """

    @staticmethod
    def add_trend_chart(
        worksheet: Worksheet,
    ) -> None:
        last_row = 3

        while (
            worksheet.cell(
                row=last_row + 1,
                column=1,
            ).value
            is not None
        ):
            last_row += 1

        categories = Reference(
            worksheet,
            min_col=1,
            min_row=4,
            max_row=last_row,
        )

        data = Reference(
            worksheet,
            min_col=8,
            min_row=3,
            max_row=last_row,
        )

        chart = LineChart()

        chart.title = "Aylık Uzlaştırma Oranı"
        chart.style = 10

        chart.height = CHART_HEIGHT
        chart.width = CHART_WIDTH

        chart.y_axis.title = "%"
        chart.x_axis.title = "Ay"

        chart.legend = None
        chart.y_axis.majorGridlines = ChartLines()

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(
            categories,
        )

        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100

        worksheet.add_chart(
            chart,
            TREND_CHART_POSITION,
        )

    @staticmethod
    def add_supplier_chart(
        worksheet: Worksheet,
    ) -> None:
        last_row = 3

        while (
            worksheet.cell(
                row=last_row + 1,
                column=10,
            ).value
            is not None
        ):
            last_row += 1

        categories = Reference(
            worksheet,
            min_col=10,
            min_row=4,
            max_row=last_row,
        )

        data = Reference(
            worksheet,
            min_col=11,
            min_row=3,
            max_row=last_row,
        )

        chart = BarChart()

        chart.type = "bar"
        chart.style = 10

        chart.title = "En Fazla Hareket Yapılan Tedarikçiler"

        chart.height = CHART_HEIGHT
        chart.width = CHART_WIDTH

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(
            categories,
        )

        worksheet.add_chart(
            chart,
            SUPPLIER_CHART_POSITION,
        )

    @staticmethod
    def add_warehouse_chart(
        worksheet: Worksheet,
    ) -> None:
        last_row = 3

        while (
            worksheet.cell(
                row=last_row + 1,
                column=13,
            ).value
            is not None
        ):
            last_row += 1

        categories = Reference(
            worksheet,
            min_col=13,
            min_row=4,
            max_row=last_row,
        )

        data = Reference(
            worksheet,
            min_col=14,
            min_row=3,
            max_row=last_row,
        )

        chart = BarChart()

        chart.type = "bar"
        chart.style = 10

        chart.title = "Ambar Bazlı Hareket Dağılımı"

        chart.height = CHART_HEIGHT
        chart.width = CHART_WIDTH

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(
            categories,
        )

        worksheet.add_chart(
            chart,
            WAREHOUSE_CHART_POSITION,
        )

    @staticmethod
    def add_top_difference_chart(
        worksheet: Worksheet,
    ) -> None:
        last_row = 3

        while (
            worksheet.cell(
                row=last_row + 1,
                column=16,
            ).value
            is not None
        ):
            last_row += 1

        categories = Reference(
            worksheet,
            min_col=16,
            min_row=4,
            max_row=last_row,
        )

        data = Reference(
            worksheet,
            min_col=17,
            min_row=3,
            max_row=last_row,
        )

        chart = BarChart()

        chart.type = "bar"
        chart.style = 10

        chart.title = "En Büyük Tutar Farkları"

        chart.height = CHART_HEIGHT
        chart.width = CHART_WIDTH

        chart.add_data(
            data,
            titles_from_data=True,
        )

        chart.set_categories(
            categories,
        )

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        chart.series[0].graphicalProperties = GraphicalProperties(
            solidFill="4472C4",
        )

        worksheet.add_chart(
            chart,
            TOP_DIFFERENCE_CHART_POSITION,
        )
