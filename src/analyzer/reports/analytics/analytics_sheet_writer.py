from __future__ import annotations

from typing import cast

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from analyzer.reports.dashboard_summary import DashboardSummary

from ..page_setup import prepare_worksheet
from .analytics_charts import AnalyticsCharts


class AnalyticsSheetWriter:
    """
    Analytics sayfasını oluşturur.
    """

    def write(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        self._write_title(worksheet)

        self._write_trend(
            worksheet,
            summary,
        )

        self._write_suppliers(
            worksheet,
            summary,
        )

        self._write_warehouses(
            worksheet,
            summary,
        )

        self._write_top_differences(
            worksheet,
            summary,
        )

        AnalyticsCharts.add_trend_chart(worksheet)
        AnalyticsCharts.add_supplier_chart(worksheet)
        AnalyticsCharts.add_warehouse_chart(worksheet)
        AnalyticsCharts.add_top_difference_chart(worksheet)

        prepare_worksheet(worksheet)

    def _write_title(
        self,
        worksheet: Worksheet,
    ) -> None:
        worksheet["A1"] = "Analytics"

    def _write_trend(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        worksheet["A3"] = "Ay"
        worksheet["B3"] = "Toplam"
        worksheet["C3"] = "Eşleşen"
        worksheet["D3"] = "MKYS Eksik"
        worksheet["E3"] = "TDMS Eksik"
        worksheet["F3"] = "Tutar Farkı"
        worksheet["G3"] = "Tüketim Farkı"
        worksheet["H3"] = "Uzlaştırma %"

        row = 4

        for month in summary.trend.months:
            cell = cast(Cell, worksheet.cell(row=row, column=1))
            cell.value = month.label

            cell = cast(Cell, worksheet.cell(row=row, column=2))
            cell.value = month.total_records

            cell = cast(Cell, worksheet.cell(row=row, column=3))
            cell.value = month.matched_records

            cell = cast(Cell, worksheet.cell(row=row, column=4))
            cell.value = month.missing_in_mkys

            cell = cast(Cell, worksheet.cell(row=row, column=5))
            cell.value = month.missing_in_tdms

            cell = cast(Cell, worksheet.cell(row=row, column=6))
            cell.value = month.amount_difference_count

            cell = cast(Cell, worksheet.cell(row=row, column=7))
            cell.value = month.consumption_difference_count

            rate_cell = cast(
                Cell,
                worksheet.cell(row=row, column=8),
            )

            rate_cell.value = month.match_rate / 100
            rate_cell.number_format = "0.0%"

            row += 1

    def _write_suppliers(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        worksheet["J1"] = "Tedarikçi Analizi"
        worksheet["J3"] = "Tedarikçi"
        worksheet["K3"] = "Kayıt"

        row = 4

        for supplier in summary.suppliers:
            cell = cast(Cell, worksheet.cell(row=row, column=10))
            cell.value = supplier.supplier

            cell = cast(Cell, worksheet.cell(row=row, column=11))
            cell.value = supplier.total_records

            row += 1

    def _write_warehouses(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        worksheet["M1"] = "Ambar Analizi"
        worksheet["M3"] = "Ambar"
        worksheet["N3"] = "Kayıt"

        row = 4

        for warehouse in summary.warehouses:
            cell = cast(Cell, worksheet.cell(row=row, column=13))
            cell.value = warehouse.warehouse

            cell = cast(Cell, worksheet.cell(row=row, column=14))
            cell.value = warehouse.total_records

            row += 1

    def _write_top_differences(
        self,
        worksheet: Worksheet,
        summary: DashboardSummary,
    ) -> None:
        worksheet["P1"] = "En Büyük Tutar Farkları"
        worksheet["P3"] = "Stok"
        worksheet["Q3"] = "Fark"

        row = 4

        for difference in summary.top_differences:
            cell = cast(Cell, worksheet.cell(row=row, column=16))
            cell.value = difference.stock_name

            cell = cast(Cell, worksheet.cell(row=row, column=17))
            cell.value = difference.difference

            row += 1
