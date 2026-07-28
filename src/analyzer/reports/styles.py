from __future__ import annotations

from openpyxl.cell import Cell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import CellIsRule
from decimal import Decimal

#
# Fonts
#

TITLE_FONT = Font(
    bold=True,
    size=16,
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

NORMAL_FONT = Font(
    bold=False,
)

SUMMARY_TITLE_FONT = Font(
    bold=True,
    size=12,
    color="1F1F1F",
)

SUMMARY_TEXT_FONT = Font(
    size=10,
    color="404040",
)

#
# Fill
#

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="4472C4",
)

SUMMARY_FILL = PatternFill(
    fill_type="solid",
    fgColor="F3F6FA",
)

#
# Borders
#

THIN_SIDE = Side(
    border_style="thin",
    color="000000",
)

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

SUMMARY_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

#
# Alignments
#

CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
)

LEFT_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="center",
)

RIGHT_ALIGNMENT = Alignment(
    horizontal="right",
    vertical="center",
)

#
# Number formats
#

AMOUNT_FORMAT = '#,##0.00'

INTEGER_FORMAT = '#,##0'

DATE_FORMAT = 'DD.MM.YYYY'

#
# Difference fills
#

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

# Zebra satırlar
ZEBRA_FILL = PatternFill(
    fill_type="solid",
    fgColor="F8F9FA",
)

# Negatif değerler
NEGATIVE_FONT = Font(
    color="C00000",
)

# Pozitif değerler
POSITIVE_FONT = Font(
    color="006100",
)


def apply_title(
    cell: Cell,
) -> None:
    """
    Başlık hücresini biçimlendirir.
    """

    cell.font = TITLE_FONT
    cell.alignment = LEFT_ALIGNMENT


def apply_header(
    cell: Cell,
) -> None:
    """
    Tablo başlıklarını biçimlendirir.
    """

    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGNMENT


def apply_text(
    cell: Cell,
) -> None:
    """
    Normal metin hücresi.
    """

    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = LEFT_ALIGNMENT


def apply_date(
    cell: Cell,
) -> None:
    """
    Tarih hücresi.
    """

    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGNMENT
    cell.number_format = DATE_FORMAT


def apply_decimal(
    cell: Cell,
) -> None:
    """
    Ondalıklı sayı hücresi.
    """

    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = RIGHT_ALIGNMENT
    cell.number_format = AMOUNT_FORMAT


def apply_integer(
    cell: Cell,
) -> None:
    """
    Tam sayı hücresi.
    """

    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = RIGHT_ALIGNMENT
    cell.number_format = INTEGER_FORMAT


def apply_currency(
    cell: Cell,
) -> None:
    """
    Para tutarı hücresi.
    """

    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = RIGHT_ALIGNMENT
    cell.number_format = AMOUNT_FORMAT


def format_worksheet(
    worksheet: Worksheet,
) -> None:
    """
    Sayfa genel ayarlarını uygular.
    """

    worksheet.freeze_panes = "A4"

    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:

        length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            if cell.value is None:
                continue

            length = max(
                length,
                len(str(cell.value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(length + 2, 12),
            50,
        )


def apply_difference_rules(
    worksheet,
    first_row: int,
    last_row: int,
    column: int,
) -> None:
    """
    Difference sütununa koşullu biçimlendirme uygular.

    0      -> Yeşil
    >0     -> Sarı
    <0     -> Kırmızı
    """

    if last_row < first_row:
        return

    from openpyxl.utils import get_column_letter

    column_letter = get_column_letter(column)

    cell_range = (
        f"{column_letter}{first_row}:"
        f"{column_letter}{last_row}"
    )

    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=["0"],
            fill=GREEN_FILL,
        ),
    )

    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=YELLOW_FILL,
        ),
    )

    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=RED_FILL,
        ),
    )


def apply_zebra_rows(
    worksheet: Worksheet,
    first_row: int,
    last_row: int,
) -> None:
    """
    Alternatif satır renklendirmesi uygular.
    """
    for row in range(first_row, last_row + 1):

        if row % 2 == 0:

            for cell in worksheet[row]:
                cell.fill = ZEBRA_FILL


def apply_number_format(
    worksheet: Worksheet,
    first_row: int,
    last_row: int,
    columns: list[int],
) -> None:
    for row in range(first_row, last_row + 1):

        for column in columns:

            cell = worksheet.cell(
                row=row,
                column=column,
            )

            cell.number_format = AMOUNT_FORMAT


def apply_negative_font(
    worksheet: Worksheet,
    first_row: int,
    last_row: int,
    column: int,
) -> None:
    for row in range(first_row, last_row + 1):

        cell = worksheet.cell(
            row=row,
            column=column,
        )

        if isinstance(cell.value, (int, float, Decimal)) and cell.value < 0:
            cell.font = NEGATIVE_FONT


def apply_kpi_title(
    cell: Cell,
) -> None:
    """
    KPI kartı başlığını biçimlendirir.
    """

    cell.font = Font(
        bold=True,
        size=11,
        color="666666",
    )

    cell.alignment = CENTER_ALIGNMENT

    cell.fill = PatternFill(
        fill_type="solid",
        fgColor="F2F2F2",
    )

    cell.border = THIN_BORDER


def apply_kpi_value(
    cell: Cell,
) -> None:
    """
    KPI kartı değerini biçimlendirir.
    """

    cell.font = Font(
        bold=True,
        size=22,
        color="1F4E78",
    )

    cell.alignment = CENTER_ALIGNMENT

    cell.border = THIN_BORDER

    cell.number_format = INTEGER_FORMAT


def apply_kpi_card(
    worksheet: Worksheet,
    first_row: int,
    last_row: int,
    first_column: int,
    last_column: int,
) -> None:
    """
    KPI kartının tamamına ortak görünüm uygular.
    """

    card_fill = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF",
    )

    for row in range(first_row, last_row + 1):

        for column in range(
            first_column,
            last_column + 1,
        ):

            cell = worksheet.cell(
                row=row,
                column=column,
            )

            cell.fill = card_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT


def apply_summary_title(
    cell: Cell,
) -> None:
    """
    Executive Summary başlığını biçimlendirir.
    """

    cell.font = SUMMARY_TITLE_FONT
    cell.alignment = LEFT_ALIGNMENT


def apply_summary_text(
    cell: Cell,
) -> None:
    """
    Executive Summary metnini biçimlendirir.
    """

    cell.font = SUMMARY_TEXT_FONT
    cell.alignment = LEFT_ALIGNMENT


def apply_summary_box(
    worksheet: Worksheet,
    first_row: int,
    last_row: int,
    first_column: int,
    last_column: int,
) -> None:
    """
    Executive Summary kutusunu biçimlendirir.
    """

    for row in range(first_row, last_row + 1):

        for column in range(first_column, last_column + 1):

            cell = worksheet.cell(
                row=row,
                column=column,
            )

            cell.fill = SUMMARY_FILL
            cell.border = SUMMARY_BORDER