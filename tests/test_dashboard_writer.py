from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
)

from analyzer.reports.dashboard_summary import DashboardSummary
from analyzer.reports.dashboard_writer import DashboardWriter

from tests.helpers.reconciliation_result_factory import (
    reconciliation_result,
)


def build_summary() -> DashboardSummary:

    return DashboardSummary.from_result(
        reconciliation_result(),
    )


def test_dashboard_title() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    assert (
        worksheet["A1"].value
        == "MKYS - TDMS Uzlaştırma Dashboard"
    )


def test_dashboard_kpi_titles() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    assert worksheet["A8"].value == "📊 Toplam Kayıt"
    assert worksheet["D8"].value == "✔ Eşleşen Kayıt"
    assert worksheet["G8"].value == "⚠ MKYS Eksik"
    assert worksheet["J8"].value == "⚠ TDMS Eksik"

    assert worksheet["A12"].value == "💰 Tutar Farkı"
    assert worksheet["D12"].value == "📦 Tüketim Farkı"


def test_dashboard_kpi_values() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    summary = build_summary()

    DashboardWriter().write_dashboard(
        worksheet,
        summary,
    )

    assert worksheet["A9"].value == summary.total_records
    assert worksheet["D9"].value == summary.matched_records
    assert worksheet["G9"].value == summary.missing_in_mkys
    assert worksheet["J9"].value == summary.missing_in_tdms

    assert worksheet["A13"].value == summary.amount_difference_count
    assert worksheet["D13"].value == summary.consumption_difference_count


def test_dashboard_prepared() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    assert worksheet.freeze_panes == "A4"

    assert worksheet.auto_filter.ref is not None


def test_dashboard_contains_pie_chart() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    charts = worksheet._charts

    assert len(charts) == 6

    assert isinstance(
        charts[0],
        PieChart,
    )

    assert charts[0].title is not None


def test_dashboard_contains_bar_chart() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    charts = worksheet._charts

    assert isinstance(
        charts[1],
        BarChart,
    )

    assert charts[1].title is not None


def test_dashboard_chart_types() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    charts = worksheet._charts

    assert len(charts) == 6

    assert isinstance(charts[0], PieChart)
    assert isinstance(charts[1], BarChart)
    assert isinstance(charts[2], LineChart)
    assert isinstance(charts[3], BarChart)


def test_dashboard_chart_titles() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    charts = worksheet._charts

    assert len(charts) == 6

    assert (
        charts[0].title.tx.rich.p[0].r[0].t
        == "Giriş Hareketleri"
    )

    assert (
        charts[1].title.tx.rich.p[0].r[0].t
        == "MKYS - TDMS Fark Analizi"
    )

    assert (
        charts[2].title.tx.rich.p[0].r[0].t
        == "Aylık Uzlaştırma Oranı"
    )

    assert (
        charts[3].title.tx.rich.p[0].r[0].t
        == "En Aktif Tedarikçiler"
    )


def test_dashboard_executive_summary() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    assert worksheet["A2"].value.endswith(
        "Executive Summary",
    )

    assert worksheet["A3"].value.startswith(
        "Uzlaştırma Oranı",
    )

    assert (
        "kayıt başarıyla eşleşti"
        in worksheet["A4"].value
    )


def test_dashboard_quality_progress() -> None:

    workbook = Workbook()

    worksheet = workbook.active

    DashboardWriter().write_dashboard(
        worksheet,
        build_summary(),
    )

    assert worksheet["J7"].value == "Kalite"

    assert worksheet["K7"].value is not None

    assert worksheet["K7"].number_format == "0%"

    assert (
        len(
            worksheet.conditional_formatting,
        )
        == 1
    )