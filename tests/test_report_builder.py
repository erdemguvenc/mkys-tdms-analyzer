from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from analyzer.reconciliation.difference import (
    AmountDifference,
    ConsumptionDifference,
)
from analyzer.reconciliation.result import ReconciliationResult
from analyzer.reports.excel_report_builder import ExcelReportBuilder
from tests.helpers.movement_factory import create_movement
from tests.helpers.reconciliation_result_factory import (
    create_reconciliation_result,
)


def test_excel_report_is_created(
    tmp_path: Path,
) -> None:
    result = ReconciliationResult(
        matched=[
            create_movement(),
        ],
    )

    output = tmp_path / "report.xlsx"

    builder = ExcelReportBuilder()

    builder.build(
        result,
        output,
    )

    assert output.exists()


def test_excel_report_is_not_empty(
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"

    builder = ExcelReportBuilder()

    builder.build(
        create_reconciliation_result(),
        output,
    )

    assert output.stat().st_size > 0


def test_consumption_difference_sheet_has_conditional_formatting(
    tmp_path: Path,
) -> None:
    difference = ConsumptionDifference(
        year=2026,
        month=1,
        mkys_amount=Decimal("100"),
        tdms_amount=Decimal("80"),
    )

    result = ReconciliationResult(
        consumption_differences=[
            difference,
        ],
    )

    output = tmp_path / "report.xlsx"

    builder = ExcelReportBuilder()

    builder.build(
        result,
        output,
    )

    workbook = load_workbook(output)

    sheet = workbook["6_Tüketim_Farkları"]

    assert len(sheet.conditional_formatting) > 0


def test_amount_difference_sheet_is_created(
    tmp_path: Path,
) -> None:
    difference = AmountDifference(
        mkys=create_movement(amount=Decimal("100")),
        tdms=create_movement(amount=Decimal("120")),
    )

    result = ReconciliationResult(
        amount_differences=[
            difference,
        ],
    )

    output = tmp_path / "report.xlsx"

    builder = ExcelReportBuilder()

    builder.build(
        result,
        output,
    )

    workbook = load_workbook(output)

    assert "5_Tutar_Farkları" in workbook.sheetnames


def test_worksheet_page_setup(
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"

    builder = ExcelReportBuilder()

    builder.build(
        ReconciliationResult(
            matched=[
                create_movement(),
            ],
        ),
        output,
    )

    workbook = load_workbook(output)

    sheet = workbook["2_Giriş_Eşleşen"]

    assert sheet.freeze_panes == "A4"

    assert sheet.auto_filter.ref is not None

    assert sheet.page_setup.orientation == sheet.ORIENTATION_LANDSCAPE

    paper_size = sheet.page_setup.paperSize
    assert paper_size is not None
    assert int(paper_size) == 9

    print_title_rows = sheet.print_title_rows
    assert print_title_rows is not None

    assert (
        print_title_rows.replace(
            "$",
            "",
        )
        == "1:3"
    )


def test_dashboard_sheet_exists(
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"

    builder = ExcelReportBuilder()

    builder.build(
        ReconciliationResult(
            matched=[
                create_movement(),
            ],
        ),
        output,
    )

    workbook = load_workbook(output)

    assert "0_Dashboard" in workbook.sheetnames

    worksheet = workbook["0_Dashboard"]

    assert worksheet.page_setup.orientation == worksheet.ORIENTATION_LANDSCAPE

    assert worksheet.page_setup.fitToWidth == 1

    assert worksheet.page_setup.fitToHeight == 1

    assert worksheet.print_area is not None
