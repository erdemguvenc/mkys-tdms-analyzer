from __future__ import annotations

from typing import Any, cast

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
)
from openpyxl.worksheet.worksheet import Worksheet

from analyzer.reports.analytics.analytics_sheet_writer import AnalyticsSheetWriter
from tests.helpers.dashboard_summary_factory import create_dashboard_summary


def test_analytics_contains_four_charts() -> None:
    workbook = Workbook()

    worksheet = cast(Worksheet, workbook.active)

    AnalyticsSheetWriter().write(
        worksheet,
        create_dashboard_summary(),
    )

    charts = cast(list[Any], getattr(worksheet, "_charts"))

    assert len(charts) == 4


def test_analytics_chart_types() -> None:
    workbook = Workbook()

    worksheet = cast(Worksheet, workbook.active)

    AnalyticsSheetWriter().write(
        worksheet,
        create_dashboard_summary(),
    )

    charts = cast(list[Any], getattr(worksheet, "_charts"))

    assert isinstance(charts[0], LineChart)
    assert isinstance(charts[1], BarChart)
    assert isinstance(charts[2], BarChart)
    assert isinstance(charts[3], BarChart)


def test_analytics_chart_titles() -> None:
    workbook = Workbook()

    worksheet = cast(Worksheet, workbook.active)

    AnalyticsSheetWriter().write(
        worksheet,
        create_dashboard_summary(),
    )

    charts = cast(list[Any], getattr(worksheet, "_charts"))

    assert charts[0].title.tx.rich.p[0].r[0].t == "Aylık Uzlaştırma Oranı"

    assert (
        charts[1].title.tx.rich.p[0].r[0].t == "En Fazla Hareket Yapılan Tedarikçiler"
    )

    assert charts[2].title.tx.rich.p[0].r[0].t == "Ambar Bazlı Hareket Dağılımı"

    assert charts[3].title.tx.rich.p[0].r[0].t == "En Büyük Tutar Farkları"
